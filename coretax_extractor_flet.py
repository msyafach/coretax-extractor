#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Coretax Extractor - Modern UI with Flet
Clean, responsive, and user-friendly interface with company-based authentication
"""

import os
import re
import time
import logging
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import flet as ft
import fitz  # PyMuPDF
import pandas as pd

from db_manager import get_db
from update_ui_helper import create_update_button

def create_logo_image(width: int = 150, height: int = 50):
    """Create RSM logo image."""
    import os
    import sys
    
    # Get the correct path for both development and PyInstaller
    if getattr(sys, 'frozen', False):
        # Running as compiled executable
        base_path = sys._MEIPASS
    else:
        # Running as script
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    # Try to find the logo file
    logo_path = os.path.join(base_path, "rsm.svg")
    if not os.path.exists(logo_path):
        logo_path = os.path.join(base_path, "rsm.png")
    if not os.path.exists(logo_path):
        logo_path = "rsm.svg"  # Fallback
    
    return ft.Image(
        src=logo_path,
        width=width,
        height=height,
        fit=ft.ImageFit.CONTAIN,
        error_content=ft.Container(
            content=ft.Text("RSM", size=24, weight=ft.FontWeight.BOLD, color="#0099D8"),
            width=width,
            height=height,
            alignment=ft.alignment.center,
            bgcolor="#F5F5F5",
            border_radius=8,
        ),
    )


class AdminPanel:
    """Admin panel for managing companies."""
    
    def __init__(self, page: ft.Page, on_back):
        self.page = page
        self.on_back = on_back
        
        # RSM Colors
        self.RSM_GREY = "#5A6670"
        self.RSM_GREEN = "#2E8B3E"
        self.RSM_BLUE = "#0099D8"
        self.RSM_DARK_GREY = "#3D4449"
        
        self.db = get_db()
        self.companies = self.db.get_all_companies()
        self._build_ui()
    
    def _build_ui(self):
        """Build admin panel UI."""
        
        # Company list
        self.company_list = ft.ListView(
            spacing=8,
            padding=12,
            height=300,
        )
        
        self._refresh_company_list()
        
        # Add company fields
        self.new_company_name = ft.TextField(
            label="Company Name",
            hint_text="Enter company name",
            width=400,
        )
        
        self.new_company_npwp = ft.TextField(
            label="NPWP",
            hint_text="Enter NPWP",
            width=400,
        )
        
        # Admin panel content
        admin_content = ft.Column([
            # Header
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.ADMIN_PANEL_SETTINGS, color=self.RSM_BLUE, size=32),
                    ft.Text(
                        "Admin Panel - Manage Companies",
                        size=24,
                        weight=ft.FontWeight.BOLD,
                        color=self.RSM_GREY,
                    ),
                ], spacing=12),
                padding=ft.padding.only(bottom=24),
            ),
            
            # Company list card
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.BUSINESS, color=self.RSM_GREEN, size=20),
                            ft.Text(
                                "Company List",
                                size=16,
                                weight=ft.FontWeight.BOLD,
                                color=self.RSM_GREY,
                            ),
                        ], spacing=8),
                        
                        ft.Container(height=12),
                        
                        ft.Container(
                            content=self.company_list,
                            bgcolor=self.RSM_DARK_GREY,
                            border_radius=8,
                            padding=0,
                            border=ft.border.all(1, "#2A2E32"),
                        ),
                    ]),
                    padding=24,
                ),
                elevation=2,
            ),
            
            ft.Container(height=16),
            
            # Add company card
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.ADD_BUSINESS, color=self.RSM_BLUE, size=20),
                            ft.Text(
                                "Add New Company",
                                size=16,
                                weight=ft.FontWeight.BOLD,
                                color=self.RSM_GREY,
                            ),
                        ], spacing=8),
                        
                        ft.Container(height=16),
                        
                        self.new_company_name,
                        ft.Container(height=8),
                        self.new_company_npwp,
                        
                        ft.Container(height=16),
                        
                        ft.Row([
                            ft.ElevatedButton(
                                "Add Company",
                                icon=ft.Icons.ADD,
                                on_click=self.add_company,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=self.RSM_GREEN,
                                ),
                                height=40,
                            ),
                        ]),
                    ]),
                    padding=24,
                ),
                elevation=2,
            ),
            
            ft.Container(height=16),
            
            # Admin Settings card
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.SETTINGS, color=self.RSM_GREY, size=20),
                            ft.Text(
                                "Admin Settings",
                                size=16,
                                weight=ft.FontWeight.BOLD,
                                color=self.RSM_GREY,
                            ),
                        ], spacing=8),
                        
                        ft.Container(height=16),
                        
                        ft.Row([
                            ft.ElevatedButton(
                                "Change Username",
                                icon=ft.Icons.PERSON,
                                on_click=self.change_username,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=self.RSM_GREY,
                                ),
                                height=40,
                            ),
                            ft.ElevatedButton(
                                "Change Admin Password",
                                icon=ft.Icons.LOCK,
                                on_click=self.change_password,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=self.RSM_GREY,
                                ),
                                height=40,
                            ),
                        ], spacing=8),
                        
                        ft.Container(height=8),
                        
                        ft.Row([
                            ft.ElevatedButton(
                                "Change App Password",
                                icon=ft.Icons.VPN_KEY,
                                on_click=self.change_app_password,
                                style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=self.RSM_BLUE,
                                ),
                                height=40,
                            ),
                        ], spacing=8),
                    ]),
                    padding=24,
                ),
                elevation=2,
            ),
            
            ft.Container(height=16),
            
            # Back button
            ft.Row([
                ft.OutlinedButton(
                    "Back to Login",
                    icon=ft.Icons.ARROW_BACK,
                    on_click=lambda e: self.on_back(),
                    style=ft.ButtonStyle(
                        color=self.RSM_GREY,
                    ),
                    height=40,
                ),
            ]),
        ], scroll=ft.ScrollMode.AUTO, expand=True)
        
        self.page.controls.clear()
        self.page.add(
            ft.Container(
                content=admin_content,
                padding=20,
                expand=True,
            )
        )
        self.page.update()
    
    def _refresh_company_list(self):
        """Refresh the company list display."""
        self.company_list.controls.clear()
        
        for company_name, npwp in sorted(self.companies.items()):
            company_row = ft.Container(
                content=ft.Row([
                    ft.Column([
                        ft.Text(
                            company_name,
                            size=13,
                            weight=ft.FontWeight.W_500,
                            color=ft.Colors.WHITE,
                        ),
                        ft.Text(
                            f"NPWP: {npwp}",
                            size=11,
                            color=ft.Colors.WHITE70,
                        ),
                    ], spacing=4, expand=True),
                    
                    # Action buttons
                    ft.Row([
                        ft.IconButton(
                            icon=ft.Icons.EDIT,
                            icon_color=self.RSM_BLUE,
                            tooltip="Edit company",
                            on_click=lambda e, name=company_name, npwp_val=npwp: self.edit_company(name, npwp_val),
                        ),
                        ft.IconButton(
                            icon=ft.Icons.DELETE,
                            icon_color=ft.Colors.RED_400,
                            tooltip="Delete company",
                            on_click=lambda e, name=company_name: self.delete_company(name),
                        ),
                    ], spacing=0),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                padding=12,
                bgcolor="#2A2E32",
                border_radius=8,
            )
            
            self.company_list.controls.append(company_row)
        
        self.page.update()
    
    def add_company(self, e):
        """Add a new company."""
        company_name = self.new_company_name.value.strip()
        npwp = self.new_company_npwp.value.strip()
        
        # Add company using database
        success, message = self.db.add_company(company_name, npwp)
        
        if success:
            self.show_success(message)
            self.new_company_name.value = ""
            self.new_company_npwp.value = ""
            # Reload companies
            self.companies = self.db.get_all_companies()
            self._refresh_company_list()
        else:
            self.show_error(message)
    
    def edit_company(self, company_name: str, company_npwp: str):
        """Edit a company."""
        # Create edit fields
        edit_name_field = ft.TextField(
            label="Company Name",
            value=company_name,
            width=400,
        )
        
        edit_npwp_field = ft.TextField(
            label="NPWP",
            value=company_npwp,
            width=400,
        )
        
        error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=12,
            visible=False,
        )
        
        def save_edit(e):
            new_name = edit_name_field.value.strip()
            new_npwp = edit_npwp_field.value.strip()
            
            # Update company
            success, message = self.db.update_company(company_name, new_name, new_npwp)
            
            if success:
                dialog.open = False
                self.page.update()
                self.show_success(message)
                # Reload companies
                self.companies = self.db.get_all_companies()
                self._refresh_company_list()
            else:
                error_text.value = message
                error_text.visible = True
                self.page.update()
        
        def cancel_edit(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.EDIT, color=self.RSM_BLUE),
                ft.Text("Edit Company"),
            ]),
            content=ft.Column([
                ft.Text(f"Edit company data '{company_name}'"),
                ft.Container(height=8),
                edit_name_field,
                ft.Container(height=8),
                edit_npwp_field,
                ft.Container(height=4),
                error_text,
            ], tight=True, width=400),
            actions=[
                ft.TextButton("Cancel", on_click=cancel_edit),
                ft.ElevatedButton(
                    "Save",
                    on_click=save_edit,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=self.RSM_GREEN,
                    ),
                ),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def delete_company(self, company_name: str):
        """Delete a company."""
        def confirm_delete(e):
            dialog.open = False
            self.page.update()
            
            success, message = self.db.delete_company(company_name)
            
            if success:
                self.show_success(message)
                # Reload companies
                self.companies = self.db.get_all_companies()
                self._refresh_company_list()
            else:
                self.show_error(message)
        
        def cancel_delete(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("Confirm Delete"),
            content=ft.Text(f"Are you sure you want to delete company '{company_name}'?"),
            actions=[
                ft.TextButton("Cancel", on_click=cancel_delete),
                ft.TextButton("Delete", on_click=confirm_delete),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def show_error(self, message: str):
        """Show error message."""
        snack = ft.SnackBar(
            content=ft.Text(message, color=ft.Colors.WHITE),
            bgcolor=ft.Colors.RED_400,
        )
        self.page.overlay.append(snack)
        snack.open = True
        self.page.update()
    
    def show_success(self, message: str):
        """Show success message."""
        snack = ft.SnackBar(
            content=ft.Text(message, color=ft.Colors.WHITE),
            bgcolor=self.RSM_GREEN,
        )
        self.page.overlay.append(snack)
        snack.open = True
        self.page.update()
    
    def change_username(self, e):
        """Change admin username."""
        # Get current username
        current_username = self.db.get_admin_username()
        
        username_field = ft.TextField(
            label="New Username",
            value=current_username,
            width=300,
            autofocus=True,
            hint_text="Minimum 3 characters",
        )
        
        error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=12,
            visible=False,
        )
        
        def save_username(e):
            new_username = username_field.value.strip()
            
            if not new_username:
                error_text.value = "Username cannot be empty"
                error_text.visible = True
                self.page.update()
                return
            
            success, message = self.db.update_admin_username(new_username)
            
            if success:
                dialog.open = False
                self.page.update()
                self.show_success(message)
            else:
                error_text.value = message
                error_text.visible = True
                self.page.update()
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.PERSON, color=self.RSM_BLUE),
                ft.Text("Change Admin Username"),
            ]),
            content=ft.Column([
                ft.Text(f"Current username: {current_username}"),
                ft.Container(height=8),
                username_field,
                ft.Container(height=4),
                error_text,
            ], tight=True, width=300),
            actions=[
                ft.TextButton("Cancel", on_click=cancel),
                ft.ElevatedButton(
                    "Save",
                    on_click=save_username,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=self.RSM_GREEN,
                    ),
                ),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def change_password(self, e):
        """Change admin password."""
        current_password_field = ft.TextField(
            label="Current Password",
            password=True,
            can_reveal_password=True,
            width=300,
            autofocus=True,
        )
        
        new_password_field = ft.TextField(
            label="New Password",
            password=True,
            can_reveal_password=True,
            width=300,
            hint_text="Minimum 4 characters",
        )
        
        confirm_password_field = ft.TextField(
            label="Confirm New Password",
            password=True,
            can_reveal_password=True,
            width=300,
        )
        
        error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=12,
            visible=False,
        )
        
        def save_password(e):
            current_password = current_password_field.value
            new_password = new_password_field.value
            confirm_password = confirm_password_field.value
            
            # Verify current password
            if not self.db.verify_admin_password(current_password):
                error_text.value = "Current password is incorrect"
                error_text.visible = True
                self.page.update()
                return
            
            # Check new password
            if not new_password:
                error_text.value = "New password cannot be empty"
                error_text.visible = True
                self.page.update()
                return
            
            # Check confirmation
            if new_password != confirm_password:
                error_text.value = "Password confirmation does not match"
                error_text.visible = True
                self.page.update()
                return
            
            success, message = self.db.update_admin_password(new_password)
            
            if success:
                dialog.open = False
                self.page.update()
                self.show_success(message)
            else:
                error_text.value = message
                error_text.visible = True
                self.page.update()
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.LOCK, color=self.RSM_BLUE),
                ft.Text("Change Admin Password"),
            ]),
            content=ft.Column([
                current_password_field,
                ft.Container(height=8),
                new_password_field,
                ft.Container(height=8),
                confirm_password_field,
                ft.Container(height=4),
                error_text,
            ], tight=True, width=300),
            actions=[
                ft.TextButton("Cancel", on_click=cancel),
                ft.ElevatedButton(
                    "Save",
                    on_click=save_password,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=self.RSM_GREEN,
                    ),
                ),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def change_app_password(self, e):
        """Change application password."""
        current_password_field = ft.TextField(
            label="Current App Password",
            password=True,
            can_reveal_password=True,
            width=300,
            autofocus=True,
        )
        
        new_password_field = ft.TextField(
            label="New App Password",
            password=True,
            can_reveal_password=True,
            width=300,
            hint_text="Minimum 4 characters",
        )
        
        confirm_password_field = ft.TextField(
            label="Confirm New Password",
            password=True,
            can_reveal_password=True,
            width=300,
        )
        
        error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=12,
            visible=False,
        )
        
        def save_app_password(e):
            current_password = current_password_field.value
            new_password = new_password_field.value
            confirm_password = confirm_password_field.value
            
            # Verify current password
            if not self.db.verify_app_password(current_password):
                error_text.value = "Current app password is incorrect"
                error_text.visible = True
                self.page.update()
                return
            
            # Check new password
            if not new_password:
                error_text.value = "New password cannot be empty"
                error_text.visible = True
                self.page.update()
                return
            
            # Check confirmation
            if new_password != confirm_password:
                error_text.value = "Password confirmation does not match"
                error_text.visible = True
                self.page.update()
                return
            
            success, message = self.db.update_app_password(new_password)
            
            if success:
                dialog.open = False
                self.page.update()
                self.show_success(message)
            else:
                error_text.value = message
                error_text.visible = True
                self.page.update()
        
        def cancel(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.VPN_KEY, color=self.RSM_BLUE),
                ft.Text("Change App Password"),
            ]),
            content=ft.Column([
                ft.Text("This password is used when first opening the application"),
                ft.Container(height=8),
                current_password_field,
                ft.Container(height=8),
                new_password_field,
                ft.Container(height=8),
                confirm_password_field,
                ft.Container(height=4),
                error_text,
            ], tight=True, width=300),
            actions=[
                ft.TextButton("Cancel", on_click=cancel),
                ft.ElevatedButton(
                    "Save",
                    on_click=save_app_password,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=self.RSM_GREEN,
                    ),
                ),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()


class LoginPage:
    """Login page with company selection."""
    
    def __init__(self, page: ft.Page, on_login_success):
        self.page = page
        self.on_login_success = on_login_success
        
        # RSM Colors
        self.RSM_GREY = "#5A6670"
        self.RSM_GREEN = "#2E8B3E"
        self.RSM_BLUE = "#0099D8"
        self.RSM_DARK_GREY = "#3D4449"
        
        self.selected_company = None
        self._build_ui()
    
    def _build_ui(self):
        """Build login UI."""
        
        # Load companies from database
        db = get_db()
        companies = db.get_all_companies()
        
        # Company dropdown
        self.company_dropdown = ft.Dropdown(
            label="Select Company",
            hint_text="Choose your company",
            options=[ft.dropdown.Option(company) for company in sorted(companies.keys())],
            width=400,
            on_change=self.on_company_selected,
            border_color=self.RSM_BLUE,
            focused_border_color=self.RSM_GREEN,
        )
        
        # Login button
        self.login_button = ft.ElevatedButton(
            "Login",
            icon=ft.Icons.LOGIN,
            on_click=self.handle_login,
            disabled=True,
            style=ft.ButtonStyle(
                color=ft.Colors.WHITE,
                bgcolor=self.RSM_GREEN,
            ),
            height=48,
            width=200,
        )
        
        # Error message
        self.error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=13,
            visible=False,
        )
        
        # Login card
        login_card = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    # RSM Logo
                    ft.Container(
                        content=create_logo_image(width=150, height=50),
                        alignment=ft.alignment.center,
                    ),
                    
                    ft.Container(height=24),
                    
                    # Title
                    ft.Text(
                        "Coretax PDF Extractor",
                        size=28,
                        weight=ft.FontWeight.BOLD,
                        color=self.RSM_GREY,
                        text_align=ft.TextAlign.CENTER,
                    ),
                    
                    ft.Container(height=8),
                    
                    # Subtitle
                    ft.Text(
                        "Please select your company to continue",
                        size=14,
                        color=self.RSM_GREY,
                        text_align=ft.TextAlign.CENTER,
                    ),
                    
                    ft.Container(height=32),
                    
                    # Company dropdown
                    self.company_dropdown,
                    
                    ft.Container(height=8),
                    
                    # Error message
                    self.error_text,
                    
                    ft.Container(height=24),
                    
                    # Login button
                    self.login_button,
                    
                    ft.Container(height=16),
                    
                    # Divider
                    ft.Divider(height=1, color="#E0E0E0"),
                    
                    ft.Container(height=8),
                    
                    # Admin button
                    ft.TextButton(
                        "Admin Login",
                        icon=ft.Icons.ADMIN_PANEL_SETTINGS,
                        on_click=self.show_admin_login,
                        style=ft.ButtonStyle(
                            color=self.RSM_GREY,
                        ),
                    ),
                    
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                padding=48,
                width=500,
            ),
            elevation=4,
        )
        
        # Center the login card
        self.page.controls.clear()
        self.page.add(
            ft.Container(
                content=login_card,
                alignment=ft.alignment.center,
                expand=True,
            )
        )
        self.page.update()
    
    def on_company_selected(self, e):
        """Handle company selection."""
        self.selected_company = e.control.value
        self.login_button.disabled = not self.selected_company
        self.error_text.visible = False
        self.page.update()
    
    def handle_login(self, e):
        """Handle login button click."""
        if not self.selected_company:
            self.error_text.value = "Please select a company"
            self.error_text.visible = True
            self.page.update()
            return
        
        # Get NPWP for selected company from database
        db = get_db()
        companies = db.get_all_companies()
        npwp = companies.get(self.selected_company)
        
        # Call success callback
        self.on_login_success(self.selected_company, npwp)
    
    def show_admin_login(self, e):
        """Show admin login dialog."""
        password_field = ft.TextField(
            label="Admin Password",
            password=True,
            can_reveal_password=True,
            width=300,
            autofocus=True,
        )
        
        error_text = ft.Text(
            "",
            color=ft.Colors.RED_400,
            size=12,
            visible=False,
        )
        
        def verify_admin(e):
            password = password_field.value
            
            if not password:
                error_text.value = "Password cannot be empty"
                error_text.visible = True
                self.page.update()
                return
            
            db = get_db()
            if db.verify_admin_password(password):
                dialog.open = False
                self.page.update()
                self.show_admin_panel()
            else:
                error_text.value = "Incorrect password"
                error_text.visible = True
                password_field.value = ""
                password_field.focus()
                self.page.update()
        
        def cancel_admin(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.ADMIN_PANEL_SETTINGS, color=self.RSM_BLUE),
                ft.Text("Admin Login"),
            ]),
            content=ft.Column([
                ft.Text("Enter admin password to continue"),
                ft.Container(height=8),
                password_field,
                ft.Container(height=4),
                error_text,
            ], tight=True),
            actions=[
                ft.TextButton("Cancel", on_click=cancel_admin),
                ft.ElevatedButton(
                    "Login",
                    on_click=verify_admin,
                    style=ft.ButtonStyle(
                        color=ft.Colors.WHITE,
                        bgcolor=self.RSM_GREEN,
                    ),
                ),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def show_admin_panel(self):
        """Show admin panel."""
        def back_to_login():
            self.page.controls.clear()
            self.page.update()
            LoginPage(self.page, self.on_login_success)
        
        AdminPanel(self.page, back_to_login)


class CoretaxExtractorApp:
    def __init__(self, page: ft.Page, company_name: str, company_npwp: str):
        self.page = page
        self.page.title = f"Coretax PDF Extractor - {company_name}"
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.page.padding = 15
        self.page.window.width = 900
        self.page.window.height = 650
        self.page.window.min_width = 800
        self.page.window.min_height = 600
        
        # Company info
        self.company_name = company_name
        self.company_npwp = company_npwp
        
        # Data
        self.pdf_files = []
        self.output_dir = ""
        self.is_processing = False
        
        # Setup logging
        self._setup_logging()
        
        # Build UI
        self._build_ui()
    
    def _setup_logging(self):
        """Configure logging for the extraction process."""
        self.log_handler = UILogHandler(self)
        
        # Get root logger and clear existing handlers
        logger = logging.getLogger()
        logger.handlers.clear()
        
        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler('coretax_extraction.log'),
                self.log_handler
            ],
            force=True
        )
    
    def _build_ui(self):
        """Build the Flet UI."""
        
        # RSM Color Scheme
        RSM_GREY = "#5A6670"
        RSM_GREEN = "#2E8B3E"
        RSM_BLUE = "#0099D8"
        RSM_DARK_GREY = "#3D4449"
        
        # Header with logout button
        header = ft.Container(
            content=ft.Column([
                # Top row with logo and logout
                ft.Row([
                    # RSM Logo
                    create_logo_image(width=120, height=40),
                    
                    # Spacer
                    ft.Container(expand=True),
                    
                    # Company info and logout
                    ft.Row([
                        ft.Icon(ft.Icons.BUSINESS, color=RSM_BLUE, size=20),
                        ft.Text(
                            self.company_name,
                            size=13,
                            weight=ft.FontWeight.W_500,
                            color=RSM_GREY,
                        ),
                        ft.Container(width=16),
                        create_update_button(self), 
                        ft.OutlinedButton(
                            "Logout",
                            icon=ft.Icons.LOGOUT,
                            on_click=self.handle_logout,
                            style=ft.ButtonStyle(
                                color=RSM_GREY,
                            ),
                            height=36,
                        ),
                    ], spacing=8),
                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                
                ft.Container(height=16),  # 16px spacing
                
                # Title
                ft.Text(
                    "Coretax PDF Extractor",
                    size=28,
                    weight=ft.FontWeight.BOLD,
                    color=RSM_GREY
                ),
                
                ft.Container(height=4),  # 4px spacing
                
                # Subtitle
                ft.Text(
                    f"Extract data from Coretax PDF files for {self.company_name}",
                    size=14,
                    color=RSM_GREY,
                    weight=ft.FontWeight.W_400,
                ),
            ], spacing=0, alignment=ft.MainAxisAlignment.START),
            padding=ft.padding.only(bottom=24)
        )
        
        # File selection section
        self.pdf_files_text = ft.Text(
            "No files selected", 
            color=RSM_GREY,
            size=13,
            italic=True
        )
        
        # File counter badge
        self.file_counter_badge = ft.Container(
            content=ft.Text(
                "0",
                size=12,
                weight=ft.FontWeight.BOLD,
                color=ft.Colors.WHITE,
            ),
            bgcolor=RSM_BLUE,
            border_radius=10,
            padding=ft.padding.symmetric(horizontal=8, vertical=2),
            visible=False,
        )
        
        file_section = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    # Header with icon, title, and counter badge
                    ft.Row([
                        ft.Icon(ft.Icons.PICTURE_AS_PDF, color=RSM_BLUE, size=24),
                        ft.Text(
                            "PDF Files", 
                            size=16, 
                            weight=ft.FontWeight.BOLD, 
                            color=RSM_GREY
                        ),
                        self.file_counter_badge,
                    ], spacing=12, alignment=ft.MainAxisAlignment.START),
                    
                    ft.Container(height=8),  # Spacing
                    
                    # Buttons row
                    ft.Row([
                        ft.ElevatedButton(
                            "Select PDF Files",
                            icon=ft.Icons.UPLOAD_FILE,
                            on_click=self.pick_pdf_files,
                            style=ft.ButtonStyle(
                                color=ft.Colors.WHITE,
                                bgcolor=RSM_BLUE,
                            ),
                            height=40,
                        ),
                        ft.OutlinedButton(
                            "Clear",
                            icon=ft.Icons.CLEAR,
                            on_click=self.clear_pdf_files,
                            style=ft.ButtonStyle(
                                color=RSM_GREY,
                            ),
                            height=40,
                        ),
                    ], spacing=8),
                    
                    ft.Container(height=8),  # Spacing
                    
                    # Status text
                    self.pdf_files_text,
                ], spacing=0, alignment=ft.MainAxisAlignment.START),
                padding=24,
            ),
            elevation=2,
        )
        
        # Output directory section
        self.output_dir_text = ft.Text(
            "No folder selected", 
            color=RSM_GREY,
            size=13,
            italic=True
        )
        
        output_section = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    # Header with icon and title
                    ft.Row([
                        ft.Icon(ft.Icons.FOLDER_OPEN, color=RSM_GREEN, size=24),
                        ft.Text(
                            "Output Folder", 
                            size=16, 
                            weight=ft.FontWeight.BOLD, 
                            color=RSM_GREY
                        ),
                    ], spacing=12, alignment=ft.MainAxisAlignment.START),
                    
                    ft.Container(height=8),  # Spacing
                    
                    # Buttons row
                    ft.Row([
                        ft.ElevatedButton(
                            "Select Output Folder",
                            icon=ft.Icons.FOLDER,
                            on_click=self.pick_output_directory,
                            style=ft.ButtonStyle(
                                color=ft.Colors.WHITE,
                                bgcolor=RSM_GREEN,
                            ),
                            height=40,
                        ),
                        ft.OutlinedButton(
                            "Clear",
                            icon=ft.Icons.CLEAR,
                            on_click=self.clear_output_directory,
                            style=ft.ButtonStyle(
                                color=RSM_GREY,
                            ),
                            height=40,
                        ),
                    ], spacing=8),
                    
                    ft.Container(height=8),  # Spacing
                    
                    # Status text
                    self.output_dir_text,
                ], spacing=0, alignment=ft.MainAxisAlignment.START),
                padding=24,
            ),
            elevation=2,
        )
        
        # Action buttons
        self.extract_button = ft.ElevatedButton(
            "Run Extraction",
            icon=ft.Icons.PLAY_ARROW,
            on_click=self.start_extraction,
            style=ft.ButtonStyle(
                color=ft.Colors.WHITE,
                bgcolor=RSM_GREEN,
            ),
            height=48,
            width=200,
        )
        
        self.clear_button = ft.OutlinedButton(
            "Clear Log",
            icon=ft.Icons.CLEAR,
            on_click=self.clear_log,
            style=ft.ButtonStyle(
                color=RSM_GREY,
            ),
            height=48,
            width=150,
        )
        
        action_section = ft.Container(
            content=ft.Row(
                [self.extract_button, self.clear_button],
                alignment=ft.MainAxisAlignment.CENTER,
                spacing=16,
            ),
            padding=ft.padding.symmetric(vertical=16),
        )
        
        # Progress bar
        self.progress_bar = ft.ProgressBar(
            width=None,
            visible=False,
            color=RSM_BLUE,
        )
        
        # Log section
        self.log_view = ft.ListView(
            spacing=4,
            padding=12,
            auto_scroll=True,
            height=220,
        )
        
        log_section = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    # Header
                    ft.Row([
                        ft.Icon(ft.Icons.TERMINAL, color=RSM_BLUE, size=24),
                        ft.Text(
                            "Extraction Log", 
                            size=16, 
                            weight=ft.FontWeight.BOLD, 
                            color=RSM_GREY
                        ),
                    ], spacing=12, alignment=ft.MainAxisAlignment.START),
                    
                    ft.Container(height=12),  # Spacing
                    
                    # Log container
                    ft.Container(
                        content=self.log_view,
                        bgcolor=ft.Colors.WHITE,
                        border_radius=8,
                        padding=0,
                        border=ft.border.all(1, "#CCCCCC"),
                    ),
                ], spacing=0),
                padding=24,
            ),
            elevation=2,
        )
        
        # Status bar
        self.status_text = ft.Text(
            "Ready to extract PDF files",
            size=13,
            color=RSM_GREY,
            weight=ft.FontWeight.W_500,
        )
        
        status_bar = ft.Container(
            content=ft.Row([
                ft.Icon(ft.Icons.INFO_OUTLINE, size=16, color=RSM_BLUE),
                self.status_text,
            ], spacing=8, alignment=ft.MainAxisAlignment.START),
            padding=16,
            bgcolor="#F5F5F5",
            border_radius=8,
            border=ft.border.all(1, "#E0E0E0"),
        )
        
        # Main layout with proper spacing (8px grid system)
        main_content = ft.Column(
            [
                header,
                ft.Container(height=8),  # 8px spacing
                
                # Cards row - equal width
                ft.Row(
                    [
                        ft.Container(
                            content=file_section,
                            expand=1,
                        ),
                        ft.Container(
                            content=output_section,
                            expand=1,
                        ),
                    ],
                    spacing=16,  # 16px spacing between cards
                ),
                
                ft.Container(height=8),  # 8px spacing
                action_section,
                
                # Progress bar container
                ft.Container(
                    content=self.progress_bar,
                    padding=ft.padding.symmetric(vertical=8),
                ),
                
                ft.Container(height=8),  # 8px spacing
                log_section,
                ft.Container(height=16),  # 16px spacing
                status_bar,
            ],
            scroll=ft.ScrollMode.AUTO,
            expand=True,
            spacing=0,  # Manual spacing control
        )
        
        self.page.add(main_content)
    
    def pick_pdf_files(self, e):
        """Handle PDF file selection."""
        file_picker = ft.FilePicker(on_result=self.on_pdf_files_selected)
        self.page.overlay.append(file_picker)
        self.page.update()
        
        file_picker.pick_files(
            allowed_extensions=["pdf"],
            allow_multiple=True,
            dialog_title="Select Coretax PDF files"
        )
    
    def on_pdf_files_selected(self, e: ft.FilePickerResultEvent):
        """Handle selected PDF files."""
        if e.files:
            self.pdf_files = [f.path for f in e.files]
            
            # Get file count
            file_count = len(self.pdf_files)
            
            # Update counter badge
            self.file_counter_badge.content.value = str(file_count)
            self.file_counter_badge.visible = True
            
            # Show simple count message
            if file_count == 1:
                self.pdf_files_text.value = "✓ 1 file selected"
            else:
                self.pdf_files_text.value = f"✓ {file_count} files selected"
            
            self.pdf_files_text.color = "#2E8B3E"  # RSM Green
            self.pdf_files_text.italic = False
            self.pdf_files_text.weight = ft.FontWeight.W_500
            self.page.update()
    
    def pick_output_directory(self, e):
        """Handle output directory selection."""
        dir_picker = ft.FilePicker(on_result=self.on_output_directory_selected)
        self.page.overlay.append(dir_picker)
        self.page.update()
        
        dir_picker.get_directory_path(dialog_title="Select output folder")
    
    def on_output_directory_selected(self, e: ft.FilePickerResultEvent):
        """Handle selected output directory."""
        if e.path:
            self.output_dir = e.path
            # Truncate long paths for display
            display_path = e.path if len(e.path) < 50 else "..." + e.path[-47:]
            self.output_dir_text.value = f"✓ {display_path}"
            self.output_dir_text.color = "#2E8B3E"  # RSM Green
            self.output_dir_text.italic = False
            self.output_dir_text.weight = ft.FontWeight.W_500
            self.page.update()
    
    def clear_pdf_files(self, e):
        """Clear selected PDF files."""
        self.pdf_files = []
        self.pdf_files_text.value = "No files selected"
        self.pdf_files_text.color = "#5A6670"  # RSM Grey
        self.pdf_files_text.italic = True
        self.pdf_files_text.weight = ft.FontWeight.W_400
        
        # Hide counter badge
        self.file_counter_badge.visible = False
        
        self.page.update()
        self.add_log("PDF files selection cleared", "INFO")
    
    def clear_output_directory(self, e):
        """Clear selected output directory."""
        self.output_dir = ""
        self.output_dir_text.value = "No folder selected"
        self.output_dir_text.color = "#5A6670"  # RSM Grey
        self.output_dir_text.italic = True
        self.output_dir_text.weight = ft.FontWeight.W_400
        self.page.update()
        self.add_log("Output folder selection cleared", "INFO")
    
    def add_log(self, message: str, level: str = "INFO"):
        """Add log message to the log view."""
        colors = {
            "INFO": ft.Colors.BLACK,
            "WARNING": ft.Colors.BLACK,
            "ERROR": ft.Colors.BLACK,
            "SUCCESS": ft.Colors.BLACK,
        }
        
        log_entry = ft.Text(
            message,
            size=11,
            color=colors.get(level, ft.Colors.BLACK),
            font_family="Courier New",
        )
        
        self.log_view.controls.append(log_entry)
        self.page.update()
    
    def clear_log(self, e):
        """Clear the log view."""
        self.log_view.controls.clear()
        self.page.update()
    
    def handle_logout(self, e):
        """Handle logout and return to login page."""
        def confirm_logout(e):
            dialog.open = False
            self.page.update()
            # Restart the app with login page
            self.page.controls.clear()
            self.page.update()
            show_login_page(self.page)
        
        def cancel_logout(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Text("Confirm Logout"),
            content=ft.Text("Are you sure you want to logout?"),
            actions=[
                ft.TextButton("Cancel", on_click=cancel_logout),
                ft.TextButton("Logout", on_click=confirm_logout),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    def update_status(self, message: str):
        """Update status bar."""
        self.status_text.value = message
        self.page.update()
    
    def set_processing(self, is_processing: bool):
        """Set processing state."""
        self.is_processing = is_processing
        self.extract_button.disabled = is_processing
        self.progress_bar.visible = is_processing
        self.page.update()
    
    def start_extraction(self, e):
        """Start the extraction process."""
        if not self.pdf_files:
            self.show_dialog("Missing Input", "Please select at least one PDF file.", "warning")
            return
        
        if not self.output_dir:
            self.show_dialog("Missing Output", "Please select an output folder.", "warning")
            return
        
        if self.is_processing:
            self.show_dialog("Busy", "Extraction is already running.", "info")
            return
        
        # Start extraction in background thread
        thread = threading.Thread(target=self._do_extraction, daemon=True)
        thread.start()
    
    def _do_extraction(self):
        """Background extraction process."""
        self.set_processing(True)
        self.update_status("Starting PDF extraction...")
        
        try:
            logger = logging.getLogger(__name__)
            logger.info("="*50)
            logger.info("CORETAX PDF EXTRACTION STARTED")
            logger.info("="*50)
            
            start_time = time.time()
            
            # Process all PDFs
            results, failed_files = self.process_pdf_files(self.pdf_files)
            
            total_files = len(self.pdf_files)
            successful_files = len([r for r in results if r.get('extraction_status') == 'Success'])
            incomplete_files = len([r for r in results if r.get('extraction_status') == 'Incomplete'])
            completely_failed = len(failed_files) - incomplete_files
            skipped_files = total_files - len(results) - completely_failed
            
            if not results:
                logger.warning("No data extracted from any PDF files")
                
                # Check if all files were skipped due to NPWP mismatch
                if skipped_files == total_files:
                    self.show_dialog(
                        "No Matching Files",
                        f"None of the {total_files} PDF files match your company's NPWP.\n\n"
                        f"Your company: {self.company_name}\n"
                        f"Your NPWP: {self.company_npwp}\n\n"
                        f"Please select PDF files that belong to your company.\n\n"
                        f"Check the log for details about which NPWP the files belong to.",
                        "warning"
                    )
                else:
                    self.show_dialog(
                        "Extraction Failed",
                        f"Failed to extract data from all {total_files} PDF files.\n\nPlease check the log for details.",
                        "error"
                    )
                return
            
            # Save results
            output_file = self.save_extraction_results(results, self.output_dir)
            
            # Statistics
            total_time = time.time() - start_time
            logger.info("="*50)
            logger.info("EXTRACTION COMPLETE")
            logger.info(f"Total files: {total_files}")
            logger.info(f"Successfully extracted: {successful_files}")
            logger.info(f"Incomplete extraction: {incomplete_files}")
            logger.info(f"Failed: {completely_failed}")
            logger.info(f"Skipped (NPWP mismatch): {skipped_files}")
            logger.info(f"Total time: {total_time:.2f} seconds")
            
            # Show extraction statistics
            key_fields = ['Nomor Bukti Potong', 'DPP', 'Pajak_Penghasilan', 'NPWP_NIK_Yang_Dipungut', 'Nama_Yang_Dipungut']
            for field in key_fields:
                success = sum(1 for item in results if item.get(field))
                logger.info(f"{field}: {success}/{len(results)} extracted ({(success/len(results)*100):.1f}%)")
            
            logger.info(f"Results saved to: {output_file}")
            
            # Log failed files
            if failed_files:
                logger.warning("="*50)
                logger.warning("FAILED/INCOMPLETE FILES:")
                for failed in failed_files:
                    logger.warning(f"  - {failed['filename']}: {failed['error']}")
                logger.warning("="*50)
            
            logger.info("="*50)
            
            # Prepare summary message
            summary_msg = f"EXTRACTION SUMMARY\n\n"
            summary_msg += f"Company: {self.company_name}\n"
            summary_msg += f"Total Files Processed: {total_files}\n\n"
            summary_msg += f"✓ Successfully Extracted: {successful_files}\n"
            
            if incomplete_files > 0:
                summary_msg += f"⚠ Incomplete Extraction: {incomplete_files}\n"
            
            if completely_failed > 0:
                summary_msg += f"✗ Failed: {completely_failed}\n"
            
            if skipped_files > 0:
                summary_msg += f"⊘ Skipped (NPWP Mismatch): {skipped_files}\n"
            
            summary_msg += f"\nResults saved to:\n{output_file}\n"
            
            # Show detailed failed files if any
            if failed_files:
                summary_msg += f"\n{'='*40}\n"
                summary_msg += "FAILED/INCOMPLETE FILES:\n\n"
                for i, failed in enumerate(failed_files[:10], 1):
                    summary_msg += f"{i}. {failed['filename']}\n"
                    summary_msg += f"   Error: {failed['error']}\n\n"
                
                if len(failed_files) > 10:
                    summary_msg += f"... and {len(failed_files) - 10} more files.\n"
                    summary_msg += "Check the log for complete details.\n"
            
            # Show appropriate message based on results
            if completely_failed == 0 and incomplete_files == 0:
                dialog_type = "success"
                title = "Extraction Complete ✓"
            elif completely_failed == total_files:
                dialog_type = "error"
                title = "Extraction Failed ✗"
            else:
                dialog_type = "warning"
                title = "Extraction Complete with Warnings ⚠"
            
            self.show_dialog(title, summary_msg, dialog_type)
            self.update_status(f"Extraction complete. Success: {successful_files}, Failed: {completely_failed + incomplete_files}")
            
        except Exception as e:
            logger.error(f"Extraction failed: {str(e)}")
            self.show_dialog("Extraction Error", f"An error occurred:\n\n{str(e)}", "error")
            self.update_status("Extraction failed.")
        finally:
            self.set_processing(False)
    
    def show_dialog(self, title: str, message: str, dialog_type: str = "info"):
        """Show dialog message."""
        icons = {
            "info": ft.Icons.INFO,
            "success": ft.Icons.CHECK_CIRCLE,
            "warning": ft.Icons.WARNING,
            "error": ft.Icons.ERROR,
        }
        
        colors = {
            "info": "#0099D8",     # RSM Blue
            "success": "#2E8B3E",  # RSM Green
            "warning": "#FFA500",  # Orange
            "error": "#FF4444",    # Red
        }
        
        def close_dialog(e):
            dialog.open = False
            self.page.update()
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(icons.get(dialog_type, ft.Icons.INFO), color=colors.get(dialog_type, ft.Colors.BLUE_700)),
                ft.Text(title),
            ]),
            content=ft.Text(message, selectable=True),
            actions=[
                ft.TextButton("OK", on_click=close_dialog),
            ],
        )
        
        self.page.overlay.append(dialog)
        dialog.open = True
        self.page.update()
    
    # ========================================================================
    # PDF Extraction Functions (from coretax_extractor_ui.py)
    # ========================================================================
    
    def extract_text_from_pdf(self, pdf_path: Path) -> str:
        """Extract raw text directly from PDF using PyMuPDF."""
        logger = logging.getLogger(__name__)
        
        try:
            logger.info(f"Extracting text from: {pdf_path.name}")
            
            pdf_document = fitz.open(str(pdf_path))
            full_text = ""
            page_count = pdf_document.page_count
            
            for page_num in range(page_count):
                page = pdf_document[page_num]
                page_text = page.get_text()
                full_text += page_text + "\n"
            
            pdf_document.close()
            
            logger.info(f"Extracted {len(full_text)} characters from {page_count} pages")
            return full_text.strip()
            
        except Exception as e:
            logger.error(f"Failed to extract text from {pdf_path.name}: {str(e)}")
            return ""
    
    def clean_and_normalize_pdf_text(self, text: str) -> str:
        """Clean and normalize text extracted directly from PDF."""
        if not text:
            return ""
        
        # First normalize all whitespace to single space
        cleaned = re.sub(r'\s+', ' ', text.strip())
        
        # Apply specific replacements
        replacements = {
            r'KEMENTERIAN\s*KEUANGAN': 'KEMENTERIAN KEUANGAN',
            r'BUKTI\s*PEMOTONGAN\s*DAN': 'BUKTI PEMOTONGAN DAN',
            r'PEMUNGUTAN\s*PPH': 'PEMUNGUTAN PPH',
            r'MASA\s*PAJAK': 'MASA PAJAK',
            r'TIDAK\s*FINAL': 'TIDAK FINAL',
            r'RSM\s*INDONESIA': 'RSM INDONESIA',
            r'BUKIT\s*ASAM': 'BUKIT ASAM',
            # Fix decimal/thousand separators in numbers
            r'(\d)\s+([.,])\s*(\d)': r'\1\2\3',
        }
        
        for pattern, replacement in replacements.items():
            cleaned = re.sub(pattern, replacement, cleaned, flags=re.IGNORECASE)
        
        return cleaned.strip()

    def extract_bukti_potong_fields_from_pdf(self, text: str, filename: str) -> Dict[str, str]:
        """Extract structured fields from PDF text."""
        
        data = {
            'Nomor Bukti Potong': '',
            'Masa Pajak': '',
            'NPWP_NIK_Yang_Dipungut': '',
            'Nama_Yang_Dipungut': '',
            'DPP': '',
            'Pajak_Penghasilan': '',
            'NPWP_NIK_Pemungut': '',
            'Nama_Pemungut': '',
            'Tanggal': '',
            'Jenis_Dokumen': '',
            'Nomor_Dokumen': '',
        }
        
        clean_text = self.clean_and_normalize_pdf_text(text).upper()
        
        # 1. Extract Nomor Bukti Potong
        # Karena ini output sistem dengan format konsisten, gunakan pendekatan berbasis konteks
        # Nomor Bukti Potong selalu ada setelah header "NOMOR" dan "MASA PAJAK", sebelum "A. IDENTITAS"
        bupot_patterns = [
            # Pattern 1: Antara "MASA PAJAK" dan "A. IDENTITAS" (paling robust)
            # Format: MASA PAJAK ... [NOMOR] [MM-YYYY] ... A. IDENTITAS
            r'MASA\s+PAJAK.*?([A-Z0-9]{8,10})\s+(\d{2}-\d{4}).*?A\.\s+IDENTITAS',
            
            # Pattern 2: Setelah "PEMUNGUTAN" (high confidence)
            r'PEMUNGUTAN\s+([A-Z0-9]{8,10})\s+\d{2}-\d{4}',
            
            # Pattern 3: Antara "NOMOR" dan "MASA PAJAK" dengan konteks
            r'NOMOR\s+MASA\s+PAJAK.*?([A-Z0-9]{8,10})\s+\d{2}-\d{4}',
            
            # Pattern 4: Di area header setelah "BPPU"
            r'BPPU.*?([A-Z0-9]{8,10})\s+\d{2}-\d{4}',
        ]
        
        for pattern in bupot_patterns:
            match = re.search(pattern, clean_text, re.DOTALL)
            if match:
                result = match.group(1)
                
                data['Nomor Bukti Potong'] = result
                break
        
        # 2. Extract Masa Pajak
        masa_patterns = [
            r'(\d{2}-\d{4})\s*TIDAK\s*FINAL',
            r'(\d{2}-\d{4})\s*NORMAL',
            r'MASA\s*PAJAK.*?(\d{2}-\d{4})',
            r'(\d{2}-\d{4})'
        ]
        
        for pattern in masa_patterns:
            match = re.search(pattern, clean_text)
            if match:
                masa = match.group(1)
                # Convert MM-YYYY to "Bulan YYYY" format
                if re.match(r'\d{2}-\d{4}', masa):
                    month_num, year = masa.split('-')
                    month_names = ['', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
                                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
                    try:
                        masa = f"{month_names[int(month_num)]} {year}"
                    except:
                        pass  # Keep original format if conversion fails
                data['Masa Pajak'] = masa
                break
        
        # 3. Extract NPWP numbers
        a1_match = re.search(r'A\.1\s*NPWP\s*/\s*NIK\s*:\s*(\d{15,16})', clean_text)
        if a1_match:
            data['NPWP_NIK_Yang_Dipungut'] = a1_match.group(1)
        
        c1_match = re.search(r'C\.1\s*NPWP\s*/\s*NIK\s*:\s*(\d{15,16})', clean_text)
        if c1_match:
            data['NPWP_NIK_Pemungut'] = c1_match.group(1)
        
        # 4. Extract A2 - Nama yang dipungut
        a2_match = re.search(r'A\.2\s*NAMA\s*:\s*(.*?)(?=A\.3)', clean_text, re.DOTALL)
        if a2_match:
            nama = a2_match.group(1).strip()
            nama = re.sub(r'\s+', ' ', nama)
            data['Nama_Yang_Dipungut'] = nama
        
        # 5 & 6. Extract DPP (B.5) and Pajak Penghasilan (B.7)
        # Support both inline and table formats
        
        # Try inline format first (with colon)
        b5_inline = re.search(r'B\.5\s*[^:]*:\s*(\d{1,3}(?:\.\d{3})*)', clean_text, re.IGNORECASE)
        if b5_inline:
            dpp_str = b5_inline.group(1)
            dpp_value = int(dpp_str.replace('.', ''))
            data['DPP'] = f"{dpp_value:,}"
        
        b7_inline = re.search(r'B\.7\s*[^:]*:\s*(\d{1,3}(?:\.\d{3})*)', clean_text, re.IGNORECASE)
        if b7_inline:
            tax_str = b7_inline.group(1)
            tax_value = int(tax_str.replace('.', ''))
            data['Pajak_Penghasilan'] = f"{tax_value:,}"
        
        # If not found, try table format
        # Table format: B.3 B.4 B.5 B.6 B.7 header, then data row with amounts
        if not data['DPP'] or not data['Pajak_Penghasilan']:
            table_match = re.search(r'B\.3\s+B\.4\s+B\.5\s+B\.6\s+B\.7(.*?)B\.8', clean_text, re.DOTALL | re.IGNORECASE)
            if table_match:
                table_content = table_match.group(1)
                # Find all numbers with thousand separators in table
                table_numbers = re.findall(r'\d{1,3}(?:\.\d{3})+', table_content)
                
                if len(table_numbers) >= 2:
                    # First large number is DPP, last number is Tax
                    if not data['DPP']:
                        dpp_value = int(table_numbers[0].replace('.', ''))
                        data['DPP'] = f"{dpp_value:,}"
                    
                    if not data['Pajak_Penghasilan']:
                        tax_value = int(table_numbers[-1].replace('.', ''))
                        data['Pajak_Penghasilan'] = f"{tax_value:,}"
        
        # 7. Extract C3 - Nama pemungut
        c3_patterns = [
            r'C\.3\s*NAMA\s*PEMOTONG\s*DAN/ATAU\s*PEMUNGUT\s*PPh\s*:\s*(.*?)(?=C\.4)',
            r'C\.3\s*NAMA\s*:\s*(.*?)(?=C\.4)',
        ]
        
        for pattern in c3_patterns:
            c3_match = re.search(pattern, clean_text, re.DOTALL | re.IGNORECASE)
            if c3_match:
                pemungut = c3_match.group(1).strip()
                pemungut = re.sub(r'\s+', ' ', pemungut)
                data['Nama_Pemungut'] = pemungut
                break
        
        # 8. Extract C4 - Tanggal
        date_patterns = [
            r'C\.4\s*TANGGAL\s*:\s*(\d{1,2})\s+(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)\s+(\d{4})',
            r'TANGGAL\s*:\s*(\d{1,2})\s+(JANUARI|FEBRUARI|MARET|APRIL|MEI|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)\s+(\d{4})',
            r':\s*(\d{1,2})\s+(MEI|APRIL|JANUARI|FEBRUARI|MARET|JUNI|JULI|AGUSTUS|SEPTEMBER|OKTOBER|NOVEMBER|DESEMBER)\s+(\d{4})',
        ]
        
        for pattern in date_patterns:
            match = re.search(pattern, clean_text)
            if match:
                day, month, year = match.groups()
                month_map = {
                    'JANUARI': 'Januari', 'FEBRUARI': 'Februari', 'MARET': 'Maret',
                    'APRIL': 'April', 'MEI': 'Mei', 'JUNI': 'Juni',
                    'JULI': 'Juli', 'AGUSTUS': 'Agustus', 'SEPTEMBER': 'September',
                    'OKTOBER': 'Oktober', 'NOVEMBER': 'November', 'DESEMBER': 'Desember'
                }
                formatted_month = month_map.get(month.upper(), month.title())
                data['Tanggal'] = f"{day} {formatted_month} {year}"
                break
        
        # 9. Extract B8 - Jenis Dokumen
        b8_patterns = [
            r'B\.8.*?JENIS\s*DOKUMEN\s*:\s*([^\n]+?)(?=\s*TANGGAL|B\.9|$)',
            r'JENIS\s*DOKUMEN\s*:\s*([^\n]+?)(?=\s*TANGGAL|B\.9|$)',
        ]
        
        for pattern in b8_patterns:
            match = re.search(pattern, clean_text, re.DOTALL)
            if match:
                jenis_dok = match.group(1).strip()
                jenis_dok = re.sub(r'\s+', ' ', jenis_dok)
                data['Jenis_Dokumen'] = jenis_dok
                break
        
        # 10. Extract B9 - Nomor Dokumen
        # Support multiple formats:
        # Format 1 (inline): B.9 NOMOR DOKUMEN : 250331/25
        # Format 2 (multiline): B.9\nNomor Dokumen\n:\n250331/25
        b9_patterns = [
            # Format 1: Inline with space support (stops at B.10)
            r'B\.9\s*NOMOR\s*DOKUMEN\s*:\s*(.+?)(?=\s*B\.10)',
            # Format 2: Multiline format
            r'B\.9\s*\n?\s*Nomor\s*Dokumen\s*\n?\s*:\s*\n?\s*(.+?)(?=\s*B\.10)',
            # Fallback: Just NOMOR DOKUMEN (stops at B.10)
            r'NOMOR\s*DOKUMEN\s*:\s*(.+?)(?=\s*B\.10)',
        ]
        
        for pattern in b9_patterns:
            match = re.search(pattern, clean_text, re.DOTALL | re.IGNORECASE)
            if match:
                nomor_dok = match.group(1).strip()
                nomor_dok = re.sub(r'\s+', ' ', nomor_dok)
                data['Nomor_Dokumen'] = nomor_dok
                break
        
        return data
    
    def process_pdf_files(self, pdf_files: List[str]) -> tuple:
        """Process multiple PDF files and extract structured data."""
        logger = logging.getLogger(__name__)
        results = []
        failed_files = []
        skipped_files = []
        
        logger.info(f"Found {len(pdf_files)} PDF files to process")
        logger.info(f"Filtering for company: {self.company_name} (NPWP: {self.company_npwp})")
        logger.info(f"Only PDFs matching NPWP {self.company_npwp} will be processed")
        
        for i, pdf_file in enumerate(pdf_files, 1):
            try:
                pdf_path = Path(pdf_file)
                logger.info(f"Processing ({i}/{len(pdf_files)}): {pdf_path.name}")
                
                progress = (i / len(pdf_files)) * 100
                self.update_status(f"Processing... {progress:.1f}%")
                
                extracted_text = self.extract_text_from_pdf(pdf_path)
                
                if not extracted_text:
                    error_msg = "No text extracted from PDF"
                    logger.warning(f"{error_msg}: {pdf_path.name}")
                    failed_files.append({
                        'filename': pdf_path.name,
                        'error': error_msg
                    })
                    continue
                
                structured_data = self.extract_bukti_potong_fields_from_pdf(extracted_text, pdf_path.name)
                structured_data['source_file'] = pdf_path.name
                
                # Check if this PDF belongs to the logged-in company (using NPWP ONLY)
                npwp_dipungut = structured_data.get('NPWP_NIK_Yang_Dipungut', '').strip()
                nama_dipungut = structured_data.get('Nama_Yang_Dipungut', '').strip()
                
                # Clean NPWP for comparison (remove dots, dashes, spaces)
                def clean_npwp(npwp: str) -> str:
                    return ''.join(c for c in npwp if c.isalnum())
                
                company_npwp_clean = clean_npwp(self.company_npwp)
                pdf_npwp_clean = clean_npwp(npwp_dipungut)
                
                # Compare NPWP ONLY (if both exist and not empty)
                if company_npwp_clean and pdf_npwp_clean:
                    if company_npwp_clean != pdf_npwp_clean:
                        # This PDF doesn't belong to the logged-in company (based on NPWP)
                        # Display both name and NPWP for clarity in logging
                        display_info = f"{nama_dipungut} (NPWP: {npwp_dipungut})" if nama_dipungut else f"NPWP: {npwp_dipungut}"
                        logger.warning(f"Skipping {pdf_path.name}: Belongs to {display_info}")
                        logger.warning(f"  Expected NPWP: {self.company_npwp} ({self.company_name})")
                        logger.warning(f"  Found NPWP: {npwp_dipungut}")
                        skipped_files.append({
                            'filename': pdf_path.name,
                            'company_name': nama_dipungut,
                            'company_npwp': npwp_dipungut,
                            'reason': f"NPWP mismatch"
                        })
                        continue
                else:
                    # If NPWP not found, log warning but continue processing
                    logger.warning(f"{pdf_path.name}: NPWP not found in PDF or company data, processing anyway")
                
                critical_fields = ['Nomor Bukti Potong', 'DPP', 'Pajak_Penghasilan']
                missing_fields = [field for field in critical_fields if not structured_data.get(field)]
                
                if missing_fields:
                    error_msg = f"Missing critical fields: {', '.join(missing_fields)}"
                    logger.warning(f"{pdf_path.name}: {error_msg}")
                    failed_files.append({
                        'filename': pdf_path.name,
                        'error': error_msg
                    })
                    structured_data['extraction_status'] = 'Incomplete'
                else:
                    structured_data['extraction_status'] = 'Success'
                
                results.append(structured_data)
                
                logger.info(f"Processed: {pdf_path.name} - Bupot={structured_data.get('Nomor Bukti Potong', 'N/A')}, DPP={structured_data.get('DPP', 'N/A')}")
                
            except Exception as e:
                error_msg = str(e)
                logger.error(f"Failed to process {pdf_file}: {error_msg}")
                failed_files.append({
                    'filename': Path(pdf_file).name,
                    'error': error_msg
                })
        
        # Log skipped files summary
        if skipped_files:
            logger.info("="*50)
            logger.info(f"SKIPPED FILES (NPWP mismatch with {self.company_name} - NPWP: {self.company_npwp}):")
            for skipped in skipped_files:
                company_info = f"{skipped['company_name']} (NPWP: {skipped['company_npwp']})" if skipped['company_name'] else f"NPWP: {skipped['company_npwp']}"
                logger.info(f"  - {skipped['filename']}: Belongs to {company_info}")
            logger.info("="*50)
        
        return results, failed_files
    
    def save_extraction_results(self, data: List[Dict[str, str]], output_dir: str) -> str:
        """Save extracted data to Excel file with proper data types."""
        logger = logging.getLogger(__name__)
        
        try:
            column_mapping = {
                'Nomor Bukti Potong': 'Nomor Bukti Potong',
                'Masa Pajak': 'Masa Pajak',
                'NPWP_NIK_Yang_Dipungut': 'NPWP/NIK yang Dipungut',
                'Nama_Yang_Dipungut': 'Nama yang Dipungut',
                'DPP': 'DPP',
                'Pajak_Penghasilan': 'Pajak Penghasilan',
                'NPWP_NIK_Pemungut': 'NPWP/NIK Pemungut',
                'Nama_Pemungut': 'Nama Pemungut',
                'Tanggal': 'Tanggal',
                'Jenis_Dokumen': 'Jenis Dokumen',
                'Nomor_Dokumen': 'Nomor Dokumen',
                'extraction_status': 'Status',
                'source_file': 'Source File'
            }
            
            df = pd.DataFrame(data)
            df = df.rename(columns=column_mapping)
            
            # Convert data types
            logger.info("Converting data types...")
            
            # Convert DPP to integer (remove commas and convert)
            if 'DPP' in df.columns:
                df['DPP'] = df['DPP'].apply(self._convert_to_integer)
            
            # Convert Pajak Penghasilan to integer
            if 'Pajak Penghasilan' in df.columns:
                df['Pajak Penghasilan'] = df['Pajak Penghasilan'].apply(self._convert_to_integer)
            
            # Convert Tanggal to datetime
            if 'Tanggal' in df.columns:
                df['Tanggal'] = df['Tanggal'].apply(self._convert_to_date)
            
            # Ensure string types for text fields
            string_columns = [
                'Nomor Bukti Potong', 'Masa Pajak', 
                'NPWP/NIK yang Dipungut', 'Nama yang Dipungut',
                'NPWP/NIK Pemungut', 'Nama Pemungut',
                'Jenis Dokumen', 'Nomor Dokumen',
                'Status', 'Source File'
            ]
            
            for col in string_columns:
                if col in df.columns:
                    df[col] = df[col].astype(str).replace('nan', '')
            
            # Create filename with company name and timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Sanitize company name for filename (remove special characters)
            safe_company_name = "".join(c for c in self.company_name if c.isalnum() or c in (' ', '-', '_')).strip()
            safe_company_name = safe_company_name.replace(' ', '_')
            
            output_file = os.path.join(output_dir, f"coretax_{safe_company_name}_{timestamp}.xlsx")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Coretax_Extraction', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Coretax_Extraction']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Format columns
                from openpyxl.styles import numbers
                
                for col_idx, col in enumerate(df.columns, 1):
                    # Format NPWP columns as TEXT to preserve leading zeros
                    if 'NPWP' in col or 'NIK' in col:
                        for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value:
                                # Force text format
                                cell.number_format = '@'
                                # Ensure value is string
                                cell.value = str(cell.value)
                    
                    # Format number columns
                    elif col in ['DPP', 'Pajak Penghasilan']:
                        for row_idx in range(2, len(df) + 2):  # Start from row 2 (after header)
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'  # Thousand separator format
                    
                    # Format date column
                    elif col == 'Tanggal':
                        for row_idx in range(2, len(df) + 2):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            if cell.value:
                                cell.number_format = 'DD MMM YYYY'  # Date format
            
            logger.info(f"Results saved to: {output_file}")
            logger.info(f"Data types applied: NPWP (text), DPP (integer), Pajak (integer), Tanggal (date)")
            return output_file
            
        except Exception as e:
            logger.error(f"Failed to save results: {str(e)}")
            raise
    
    def _convert_to_integer(self, value: str) -> int:
        """Convert string with commas to integer."""
        if not value or value == '' or value == 'nan':
            return None
        try:
            # Remove commas and convert to int
            cleaned = str(value).replace(',', '').replace('.', '').strip()
            return int(cleaned) if cleaned else None
        except (ValueError, AttributeError):
            return None
    
    def _convert_to_date(self, value: str):
        """Convert Indonesian date string to datetime."""
        if not value or value == '' or value == 'nan':
            return None
        
        try:
            # Parse Indonesian date format: "5 Juni 2025"
            month_map = {
                'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4,
                'Mei': 5, 'Juni': 6, 'Juli': 7, 'Agustus': 8,
                'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
            }
            
            parts = str(value).strip().split()
            if len(parts) == 3:
                day = int(parts[0])
                month = month_map.get(parts[1], 1)
                year = int(parts[2])
                return datetime(year, month, day)
            
            return None
        except (ValueError, AttributeError, IndexError):
            return None


class UILogHandler(logging.Handler):
    """Custom logging handler to display logs in the UI."""
    
    def __init__(self, app_instance):
        super().__init__()
        self.app = app_instance
        
    def emit(self, record):
        try:
            msg = self.format(record)
            level = record.levelname
            self.app.add_log(msg, level)
        except Exception:
            pass


def show_login_page(page: ft.Page):
    """Show login page."""
    def on_login_success(company_name: str, company_npwp: str):
        # Clear login page and show main app
        page.controls.clear()
        page.update()
        CoretaxExtractorApp(page, company_name, company_npwp)
    
    LoginPage(page, on_login_success)


def show_splash_screen(page: ft.Page):
    """Show splash screen with logo."""
    # RSM Colors
    RSM_GREY = "#5A6670"
    RSM_GREEN = "#2E8B3E"
    RSM_BLUE = "#0099D8"
    
    # Logo with animation
    logo = ft.Container(
        content=create_logo_image(width=250, height=83),
        alignment=ft.alignment.center,
        animate_opacity=1000,
    )
    
    # Loading text
    loading_text = ft.Text(
        "Loading...",
        size=14,
        color=RSM_GREY,
        text_align=ft.TextAlign.CENTER,
        opacity=0,
        animate_opacity=500,
    )
    
    # Progress bar
    progress = ft.ProgressBar(
        width=200,
        color=RSM_BLUE,
        bgcolor="#E0E0E0",
        opacity=0,
        animate_opacity=500,
    )
    
    # Splash content
    splash_content = ft.Container(
        content=ft.Column([
            ft.Container(expand=True),  # Top spacer
            logo,
            ft.Container(height=40),
            loading_text,
            ft.Container(height=16),
            progress,
            ft.Container(expand=True),  # Bottom spacer
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        alignment=ft.alignment.center,
        expand=True,
    )
    
    page.controls.clear()
    page.add(splash_content)
    page.update()
    
    # Animate loading text and progress
    import time
    time.sleep(0.3)
    loading_text.opacity = 1
    progress.opacity = 1
    page.update()
    
    # Simulate loading (initialize database, etc.)
    time.sleep(1.5)
    
    # Transition to password screen
    show_password_screen(page)


def show_password_screen(page: ft.Page):
    """Show initial password screen."""
    # RSM Colors
    RSM_GREY = "#5A6670"
    RSM_GREEN = "#2E8B3E"
    RSM_BLUE = "#0099D8"
    
    # Get database instance
    db = get_db()
    
    password_field = ft.TextField(
        label="Password",
        password=True,
        can_reveal_password=True,
        width=300,
        autofocus=True,
        hint_text="Enter application password",
        on_submit=lambda e: verify_password(e),
    )
    
    error_text = ft.Text(
        "",
        color=ft.Colors.RED_400,
        size=13,
        visible=False,
    )
    
    def verify_password(e):
        password = password_field.value
        
        if not password:
            error_text.value = "Password cannot be empty"
            error_text.visible = True
            page.update()
            return
        
        if db.verify_app_password(password):
            # Password correct, show login page
            page.controls.clear()
            page.update()
            show_login_page(page)
        else:
            error_text.value = "Incorrect password!"
            error_text.visible = True
            password_field.value = ""
            password_field.focus()
            page.update()
    
    # Password screen content
    password_screen = ft.Container(
        content=ft.Column([
            ft.Container(height=20),  # Top spacing
            
            # RSM Logo (bigger)
            ft.Container(
                content=create_logo_image(width=200, height=67),
                alignment=ft.alignment.center,
            ),
            
            ft.Container(height=40),  # More spacing below logo
            
            # Title
            ft.Text(
                "Coretax PDF Extractor",
                size=32,
                weight=ft.FontWeight.BOLD,
                color=RSM_GREY,
                text_align=ft.TextAlign.CENTER,
            ),
            
            ft.Container(height=8),
            
            # Subtitle
            ft.Text(
                "Enter your password to continue",
                size=14,
                color=RSM_GREY,
                text_align=ft.TextAlign.CENTER,
            ),
            
            ft.Container(height=40),
            
            # Password field
            password_field,
            
            ft.Container(height=8),
            
            # Error message
            error_text,
            
            ft.Container(height=24),
            
            # Login button
            ft.ElevatedButton(
                "Login",
                icon=ft.Icons.LOGIN,
                on_click=verify_password,
                style=ft.ButtonStyle(
                    color=ft.Colors.WHITE,
                    bgcolor=RSM_GREEN,
                ),
                height=48,
                width=200,
            ),
            
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER),
        alignment=ft.alignment.center,
        expand=True,
    )
    
    page.controls.clear()
    page.add(password_screen)
    page.update()


def main(page: ft.Page):
    """Main entry point for Flet app."""
    page.title = "Coretax PDF Extractor"
    page.theme_mode = ft.ThemeMode.LIGHT
    page.padding = 15
    page.window.icon = "icon/app.ico"
    page.window.width = 900
    page.window.height = 650
    page.window.min_width = 800
    page.window.min_height = 600
    page.theme = ft.Theme(font_family="Roboto")
    
    # Center window on screen
    page.window.center()
    
    # Show splash screen first
    show_splash_screen(page)


if __name__ == "__main__":
    ft.app(target=main)
